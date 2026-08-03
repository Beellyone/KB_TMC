import io
import json
import re

import structlog
import xlrd
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ...auth.dependencies import get_current_user
from ...database import get_db
from ...models.act import Act, ActDate, ActStatus, AttentionMark, CostApproval, DamagedTmc, Repair
from ...models.audit_log import AuditLog
from ...models.tmc import Tmc
from ...models.user import User
from ...schemas.act import (
    ActConfirmRequest,
    ActDatesResponse,
    ActDetailResponse,
    ActResponse,
    ActUpdate,
    GuaranteeCheckRequest,
    GuaranteeCheckResponse,
    RepairResponse,
    RepairUpdate,
    UploadPreviewResponse,
    UploadPreviewRow,
    UploadValidationError,
    UploadValidationWarning,
)
from ...services.guarantee import check_guarantee

log = structlog.get_logger()
router = APIRouter(prefix="/acts", tags=["acts"])

SERIAL_PATTERN = re.compile(r"^[a-zA-Z0-9\-_.]+$")
ACTIVE_STATUSES = (ActStatus.NEW, ActStatus.CHECKING, ActStatus.DONE)


async def _audit(
    db: AsyncSession,
    user: User,
    action: str,
    entity_type: str,
    entity_id: int,
    changes: dict | None = None,
):
    entry = AuditLog(
        user_id=user.id,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        changes_json=json.dumps(changes, ensure_ascii=False, default=str) if changes else None,
    )
    db.add(entry)
    log.info("audit", action=action, entity_type=entity_type, entity_id=entity_id, user=user.username)


# ── List / Get / Update / Delete ────────────────────────────


@router.get("", response_model=list[ActResponse])
async def list_acts(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Act).order_by(Act.created_at.desc()))
    return result.scalars().all()


@router.get("/{act_id}", response_model=ActDetailResponse)
async def get_act(act_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Act).where(Act.id == act_id).options(selectinload(Act.dates))
    )
    act = result.scalar_one_or_none()
    if not act:
        raise HTTPException(status_code=404, detail="Акт не найден")
    count_result = await db.execute(
        select(func.count()).select_from(Repair).where(Repair.act_id == act_id)
    )
    count = count_result.scalar()
    try:
        resp = ActDetailResponse.model_validate(act)
        if act.dates:
            resp.dates = ActDatesResponse.model_validate(act.dates)
        else:
            resp.dates = ActDatesResponse(
                diagnostics_date=None, verification_date=None, invoice_date=None,
                return_date=None, confirmation_date=None, completion_date=None,
            )
        resp.repairs_count = count
        return resp
    except Exception as e:
        log.error("get_act_error", act_id=act_id, error=str(e))
        raise HTTPException(status_code=500, detail=str(e)) from None


@router.patch("/{act_id}", response_model=ActResponse)
async def update_act(
    act_id: int,
    body: ActUpdate,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    act = await db.get(Act, act_id)
    if not act:
        raise HTTPException(status_code=404, detail="Акт не найден")
    changes = {}
    if body.status is not None and body.status != act.status:
        changes["status"] = {"old": act.status.value, "new": body.status.value}
        act.status = body.status
    if body.attention_mark is not None and body.attention_mark != act.attention_mark:
        changes["attention_mark"] = {"old": act.attention_mark.value, "new": body.attention_mark.value}
        act.attention_mark = body.attention_mark
    if body.dates:
        result = await db.execute(
            select(ActDate).where(ActDate.act_id == act_id)
        )
        dates = result.scalar_one_or_none()
        if not dates:
            dates = ActDate(act_id=act_id)
            db.add(dates)
        for field, value in body.dates.model_dump(exclude_unset=True).items():
            if getattr(dates, field) != value:
                changes[f"dates.{field}"] = {"old": str(getattr(dates, field)), "new": str(value)}
                setattr(dates, field, value)
    if changes:
        await _audit(db, user, "update", "act", act.id, changes)
        await db.commit()
        await db.refresh(act)
    return act


@router.delete("/{act_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_act(
    act_id: int,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    act = await db.get(Act, act_id)
    if not act:
        raise HTTPException(status_code=404, detail="Акт не найден")
    await _audit(db, user, "delete", "act", act.id, {"number": act.number})
    await db.delete(act)
    await db.commit()


# ── Upload & Validate ───────────────────────────────────────


@router.post("/upload", response_model=UploadPreviewResponse)
async def upload_act(
    file: UploadFile = File(...),
    act_number: str = Form(...),
    executor_id: int = Form(...),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename or not file.filename.endswith((".xls", ".xlsx")):
        raise HTTPException(status_code=400, detail="Файл должен быть .xls или .xlsx")

    content = await file.read()
    is_xls = file.filename.endswith(".xls") and not file.filename.endswith(".xlsx")

    # Parse headers and rows into a unified format
    all_rows: list[list[str | None]] = []
    if is_xls:
        book = xlrd.open_workbook(file_contents=content)
        sheet = book.sheet_by_index(0)
        for row_idx in range(sheet.nrows):
            row_vals = []
            for col in range(sheet.ncols):
                val = sheet.cell_value(row_idx, col)
                row_vals.append(str(val).strip() if val not in ("", None) else None)
            all_rows.append(row_vals)
    else:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            raise HTTPException(status_code=400, detail="Не удалось прочитать файл")
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(cell).strip() if cell is not None else None for cell in row])
        wb.close()

    if len(all_rows) < 2:
        raise HTTPException(status_code=400, detail="Файл пуст или содержит только заголовки")

    headers = all_rows[0]
    header_map: dict[str, int] = {}
    for idx, h in enumerate(headers):
        if h:
            header_map[h.strip().lower()] = idx

    code_col = header_map.get("код")
    serial_col = header_map.get("серия номенклатуры")
    comment_col = header_map.get("комментарий")

    if code_col is None:
        raise HTTPException(status_code=400, detail="Не найдена колонка 'Код'")
    if serial_col is None:
        raise HTTPException(status_code=400, detail="Не найдена колонка 'Серия номенклатуры'")

    rows_data: list[UploadPreviewRow] = []
    errors: list[UploadValidationError] = []
    warnings: list[UploadValidationWarning] = []
    seen_keys: dict[str, int] = {}

    for row_idx, row in enumerate(all_rows[1:], start=2):
        tmc_code_raw = row[code_col] if code_col < len(row) else None
        serial_raw = row[serial_col] if serial_col < len(row) else None
        comment = row[comment_col] if comment_col is not None and comment_col < len(row) else None

        # Convert numeric codes like "10379.0" → "10379"
        tmc_code = ""
        if tmc_code_raw is not None:
            tmc_code = str(tmc_code_raw).strip()
            try:
                tmc_code = str(int(float(tmc_code)))
            except (ValueError, TypeError):
                tmc_code = str(tmc_code_raw).strip()

        serial_number = str(serial_raw).strip() if serial_raw is not None else ""
        fault_description = str(comment).strip() if comment else ""

        row_obj = UploadPreviewRow(
            row_number=row_idx,
            tmc_code=tmc_code,
            serial_number=serial_number,
            fault_description=fault_description,
        )

        # Check for duplicates within the same file
        if tmc_code and serial_number:
            dup_key = f"{tmc_code}:{serial_number}"
            if dup_key in seen_keys:
                errors.append(UploadValidationError(
                    row_number=row_idx, field="serial_number",
                    message=f"Дубликат строки {seen_keys[dup_key]}: тот же код и серийный номер",
                ))
            else:
                seen_keys[dup_key] = row_idx

        # Validate TMC code
        tmc_result = await db.execute(select(Tmc).where(Tmc.code == tmc_code))
        tmc = tmc_result.scalar_one_or_none()
        if not tmc:
            errors.append(UploadValidationError(
                row_number=row_idx, field="tmc_code",
                message=f"ТМЦ с кодом '{tmc_code}' не найден",
            ))
        else:
            row_obj.tmc_id = tmc.id
            row_obj.tmc_name = tmc.name

        # Validate serial number
        if not serial_number:
            errors.append(UploadValidationError(
                row_number=row_idx, field="serial_number",
                message="Серийный номер не указан",
            ))
        elif not SERIAL_PATTERN.match(serial_number):
            errors.append(UploadValidationError(
                row_number=row_idx, field="serial_number",
                message="Серийный номер содержит недопустимые символы (русские буквы или пробелы)",
            ))
        else:
            # Only check active acts and guarantee if TMC was found
            if tmc:
                # Check if already in active act (same tmc + serial)
                active_result = await db.execute(
                    select(DamagedTmc)
                    .join(Repair)
                    .join(Act)
                    .where(
                        DamagedTmc.tmc_id == tmc.id,
                        DamagedTmc.serial_number == serial_number,
                        Act.status.in_(ACTIVE_STATUSES),
                    )
                )
                active_damaged = active_result.scalars().first()
                if active_damaged:
                    errors.append(UploadValidationError(
                        row_number=row_idx, field="serial_number",
                        message=f"ТМЦ с серийным номером '{serial_number}' уже находится в активном акте",
                    ))

                # Check guarantee
                guarantee = await check_guarantee(db, tmc.id, serial_number)
                if guarantee.is_guarantee:
                    warnings.append(UploadValidationWarning(
                        row_number=row_idx, field="serial_number",
                        message=(
                            f"Гарантийный ремонт (исходный ремонт #{guarantee.original_repair_id}, "
                            f"осталось {guarantee.days_remaining} дн.)"
                        ),
                        is_guarantee=True,
                    ))

        rows_data.append(row_obj)

    return UploadPreviewResponse(
        rows=rows_data,
        errors=errors,
        warnings=warnings,
        file_name=file.filename or "unknown",
        executor_id=executor_id,
        act_number=act_number,
    )


# ── Confirm & Create ────────────────────────────────────────


@router.post("/confirm", response_model=ActResponse, status_code=status.HTTP_201_CREATED)
async def confirm_act(
    body: ActConfirmRequest,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    dup = await db.execute(select(Act).where(Act.number == body.number))
    if dup.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"Акт с номером '{body.number}' уже существует")

    has_attention = False
    act = Act(
        number=body.number,
        executor_id=body.executor_id,
        file_name=body.file_name,
    )
    db.add(act)
    await db.flush()

    dates = ActDate(act_id=act.id)
    db.add(dates)

    for row in body.rows:
        if not row.tmc_id:
            continue
        # Find or create DamagedTmc (unique by tmc_id + serial_number)
        result = await db.execute(
            select(DamagedTmc).where(
                DamagedTmc.tmc_id == row.tmc_id,
                DamagedTmc.serial_number == row.serial_number,
            )
        )
        damaged = result.scalar_one_or_none()
        if not damaged:
            damaged = DamagedTmc(tmc_id=row.tmc_id, serial_number=row.serial_number)
            db.add(damaged)
            await db.flush()

        # Check guarantee for cost_approval
        guarantee = await check_guarantee(db, row.tmc_id, row.serial_number, exclude_act_id=act.id)
        cost_approval = CostApproval.GUARANTEE if guarantee.is_guarantee else CostApproval.DISAPPROVED
        attention = AttentionMark.ATTENTION if guarantee.is_guarantee else AttentionMark.CALM
        if attention == AttentionMark.ATTENTION:
            has_attention = True

        repair = Repair(
            act_id=act.id,
            damaged_tmc_id=damaged.id,
            fault_description=row.fault_description or None,
            cost_approval=cost_approval,
            attention_mark=attention,
        )
        db.add(repair)

    if has_attention:
        act.attention_mark = AttentionMark.ATTENTION

    await _audit(db, user, "create", "act", act.id, {
        "number": act.number, "executor_id": act.executor_id,
        "repairs_count": len(body.rows),
    })
    await db.commit()
    await db.refresh(act)
    return act


# ── Repairs ─────────────────────────────────────────────────


@router.get("/{act_id}/repairs", response_model=list[RepairResponse])
async def list_repairs(act_id: int, db: AsyncSession = Depends(get_db)):
    if not await db.get(Act, act_id):
        raise HTTPException(status_code=404, detail="Акт не найден")
    result = await db.execute(
        select(Repair)
        .where(Repair.act_id == act_id)
        .options(
            selectinload(Repair.damaged_tmc).selectinload(DamagedTmc.repairs),
        )
    )
    repairs = result.scalars().all()
    response = []
    for r in repairs:
        dtmc = r.damaged_tmc
        tmc_result = await db.execute(select(Tmc).where(Tmc.id == dtmc.tmc_id))
        tmc = tmc_result.scalar_one_or_none()
        resp = RepairResponse.model_validate(r)
        resp.tmc_code = tmc.code if tmc else None
        resp.tmc_name = tmc.name if tmc else None
        resp.serial_number = dtmc.serial_number
        response.append(resp)
    return response


@router.patch("/{act_id}/repairs/{repair_id}", response_model=RepairResponse)
async def update_repair(
    act_id: int,
    repair_id: int,
    body: RepairUpdate,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    repair = await db.get(Repair, repair_id)
    if not repair or repair.act_id != act_id:
        raise HTTPException(status_code=404, detail="Ремонт не найден")
    changes = {}
    update_data = body.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        old = getattr(repair, field)
        old_str = old.value if hasattr(old, "value") else str(old)
        new_str = value.value if hasattr(value, "value") else str(value)
        if old_str != new_str:
            changes[field] = {"old": old_str, "new": new_str}
            setattr(repair, field, value)
    if changes:
        await _audit(db, user, "update", "repair", repair.id, changes)
        await db.commit()
        await db.refresh(repair)
    resp = RepairResponse.model_validate(repair)
    dtmc = await db.get(DamagedTmc, repair.damaged_tmc_id)
    if dtmc:
        tmc = await db.get(Tmc, dtmc.tmc_id)
        resp.tmc_code = tmc.code if tmc else None
        resp.tmc_name = tmc.name if tmc else None
        resp.serial_number = dtmc.serial_number
    return resp


# ── Guarantee Check ─────────────────────────────────────────


@router.post("/guarantee-check", response_model=GuaranteeCheckResponse)
async def guarantee_check(
    body: GuaranteeCheckRequest,
    db: AsyncSession = Depends(get_db),
):
    result = await check_guarantee(db, body.tmc_id, body.serial_number)
    return GuaranteeCheckResponse(
        is_guarantee=result.is_guarantee,
        original_repair_id=result.original_repair_id,
        original_completion_date=result.original_completion_date,
        days_remaining=result.days_remaining,
    )
